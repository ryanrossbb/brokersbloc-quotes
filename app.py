import io
import os
import json
import datetime
import traceback
import pdfplumber
import anthropic
from flask import Flask, request, send_file, jsonify, render_template_string
from openpyxl import load_workbook
from openpyxl.styles import Protection, Alignment

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

TEMPLATE_FILE = "BrokersBloc_Medical_Template.xlsx"
GROUP_COLS = [(8,9,10), (11,12,13), (14,15,16), (17,18,19), (20,21,22)]
ROW_MAP = {
    "network":12,"deductible":13,"coinsurance":14,"moop":15,
    "pcp":16,"telehealth":17,"specialist":18,"inpatient":19,
    "outpatient":20,"er":21,"urgent_care":22,"lab":23,"xray":24,
    "imaging":25,"oon_ded":27,"oon_cost":28,"oon_moop":29,
    "rx_ded":31,"rx_copays":32,"rate_ee":34,"rate_es":35,
    "rate_ec":36,"rate_ef":37,"rate_guarantee":38,
}

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BrokersBloc — Medical Quote Builder</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  :root{--bb-navy:#0B1F3F;--bb-blue:#1A56A0;--bb-light:#E8F0FF}
  body{background:#f4f6fb;font-family:'Segoe UI',sans-serif}
  .bb-header{background:var(--bb-navy);padding:18px 32px;display:flex;align-items:center;justify-content:space-between}
  .bb-header .brand{color:#fff;font-size:1.7rem;font-weight:700}
  .bb-header .site{color:#8fa8cc;font-size:.9rem}
  .page-body{max-width:900px;margin:32px auto;padding:0 16px 100px}
  .card{border:none;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px}
  .card-header{background:var(--bb-navy);color:#fff;font-weight:600;font-size:.95rem;padding:12px 20px;border-radius:10px 10px 0 0!important}
  .card-body{padding:20px 24px}
  label{font-size:.83rem;color:#445;font-weight:500;margin-bottom:3px}
  .form-control,.form-select{font-size:.9rem;border-radius:6px;border:1px solid #cdd6e8}
  .form-control:focus{border-color:var(--bb-blue);box-shadow:0 0 0 3px rgba(26,86,160,.15)}
  .subsection{font-size:.72rem;font-weight:700;letter-spacing:.8px;text-transform:uppercase;color:var(--bb-blue);margin:16px 0 8px;border-bottom:1px solid #e0e8f5;padding-bottom:4px}
  .pdf-slot{background:#f8faff;border:2px dashed #c5d5ee;border-radius:8px;padding:16px;margin-bottom:12px;transition:.2s}
  .pdf-slot:hover{border-color:var(--bb-blue);background:#f0f5ff}
  .pdf-slot.has-file{border-style:solid;border-color:#2ecc71;background:#f0fff5}
  .pdf-slot .slot-num{font-size:.75rem;font-weight:700;text-transform:uppercase;color:var(--bb-blue);letter-spacing:.5px;margin-bottom:6px}
  .pdf-status{font-size:.8rem;color:#27ae60;margin-top:4px;display:none}
  #overlay{display:none;position:fixed;inset:0;background:rgba(11,31,63,.85);z-index:9999;align-items:center;justify-content:center;flex-direction:column}
  #overlay .spinner-border{width:3rem;height:3rem;border-width:.3em}
  #overlay .msg{color:#fff;margin-top:20px;font-size:1.1rem}
  #overlay .sub{color:#8fa8cc;font-size:.85rem;margin-top:6px}
  .submit-bar{position:fixed;bottom:0;left:0;right:0;background:var(--bb-navy);padding:14px 32px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 -2px 12px rgba(0,0,0,.2);z-index:100}
  .btn-generate{background:#2ecc71;color:#fff;font-weight:600;padding:10px 32px;border-radius:6px;border:none;font-size:1rem;cursor:pointer}
  .btn-generate:hover{background:#27ae60}
  .btn-generate:disabled{background:#555;cursor:not-allowed}
  .hint{color:#8fa8cc;font-size:.85rem}
</style>
</head>
<body>
<div id="overlay">
  <div class="spinner-border text-light"></div>
  <div class="msg">Reading PDFs and extracting plan data&hellip;</div>
  <div class="sub">This usually takes 20&ndash;40 seconds per carrier</div>
</div>
<div class="bb-header">
  <span class="brand">BrokersBloc</span>
  <span class="site">brokersbloc.com</span>
</div>
<div class="page-body">
  <h4 class="mb-1 mt-1" style="color:var(--bb-navy)">Medical Quote Comparison Builder</h4>
  <p class="text-muted mb-4" style="font-size:.9rem">Upload carrier proposal PDFs — the AI reads them, extracts the plan data, and builds the populated comparison Excel automatically.</p>
  <div id="errorBox" class="alert alert-danger d-none"></div>
  <form id="quoteForm" enctype="multipart/form-data">
    <div class="card">
      <div class="card-header">Group Information</div>
      <div class="card-body">
        <div class="row g-3">
          <div class="col-md-6">
            <label>Group / Company Name *</label>
            <input type="text" class="form-control" name="group_name" placeholder="Liberty Restoration Group, Inc." required>
          </div>
          <div class="col-md-3">
            <label>Effective Date</label>
            <input type="text" class="form-control" name="effective_date" placeholder="06/01/2026">
          </div>
          <div class="col-md-3">
            <label>Quote Date</label>
            <input type="date" class="form-control" name="quote_date" id="quoteDate">
          </div>
        </div>
        <div class="subsection">Enrollment Counts</div>
        <div class="row g-3">
          <div class="col-6 col-md-3"><label>Employee Only</label><input type="number" class="form-control" name="enroll_ee" value="0" min="0"></div>
          <div class="col-6 col-md-3"><label>Emp + Spouse</label><input type="number" class="form-control" name="enroll_es" value="0" min="0"></div>
          <div class="col-6 col-md-3"><label>Emp + Child(ren)</label><input type="number" class="form-control" name="enroll_ec" value="0" min="0"></div>
          <div class="col-6 col-md-3"><label>Emp + Family</label><input type="number" class="form-control" name="enroll_ef" value="0" min="0"></div>
        </div>
        <div class="subsection">Plan Filter</div>
        <div class="row g-3">
          <div class="col-md-6">
            <label>Which plans to include from each PDF?</label>
            <input type="text" class="form-control" name="plan_filter" value="all plans" placeholder="e.g. all plans, $5,000 deductible, HDHP only">
            <div class="form-text">Examples: "all plans" &middot; "$5,000 deductible" &middot; "HDHP only"</div>
          </div>
        </div>
      </div>
    </div>
    <div class="card">
      <div class="card-header">Carrier PDFs &mdash; Upload up to 5 (one per carrier)</div>
      <div class="card-body">
        <p class="text-muted mb-3" style="font-size:.85rem">Upload each carrier's proposal PDF. The AI extracts plan names, benefits, rates, and network info automatically.</p>
        <div class="pdf-slot" id="slot0"><div class="slot-num">Carrier 1 <span class="text-success">(required)</span></div><input type="file" class="form-control" name="pdf_0" accept=".pdf" onchange="fileSelected(this,0)"><div class="pdf-status" id="status0"></div></div>
        <div class="pdf-slot" id="slot1"><div class="slot-num">Carrier 2 <span class="text-muted">(optional)</span></div><input type="file" class="form-control" name="pdf_1" accept=".pdf" onchange="fileSelected(this,1)"><div class="pdf-status" id="status1"></div></div>
        <div class="pdf-slot" id="slot2"><div class="slot-num">Carrier 3 <span class="text-muted">(optional)</span></div><input type="file" class="form-control" name="pdf_2" accept=".pdf" onchange="fileSelected(this,2)"><div class="pdf-status" id="status2"></div></div>
        <div class="pdf-slot" id="slot3"><div class="slot-num">Carrier 4 <span class="text-muted">(optional)</span></div><input type="file" class="form-control" name="pdf_3" accept=".pdf" onchange="fileSelected(this,3)"><div class="pdf-status" id="status3"></div></div>
        <div class="pdf-slot" id="slot4"><div class="slot-num">Carrier 5 <span class="text-muted">(optional)</span></div><input type="file" class="form-control" name="pdf_4" accept=".pdf" onchange="fileSelected(this,4)"><div class="pdf-status" id="status4"></div></div>
      </div>
    </div>
  </form>
</div>
<div class="submit-bar">
  <span class="hint">AI reads each PDF &middot; extracts plans &middot; builds the Excel &middot; downloads automatically</span>
  <button type="button" class="btn-generate" id="generateBtn" onclick="submitForm()">&#11015; Generate Excel</button>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script>
document.getElementById('quoteDate').valueAsDate = new Date();

function fileSelected(input, idx) {
  const slot = document.getElementById('slot'+idx);
  const status = document.getElementById('status'+idx);
  if (input.files && input.files[0]) {
    slot.classList.add('has-file');
    status.style.display = 'block';
    status.textContent = '✓ ' + input.files[0].name;
  } else {
    slot.classList.remove('has-file');
    status.style.display = 'none';
  }
}

async function submitForm() {
  const errorBox = document.getElementById('errorBox');
  errorBox.classList.add('d-none');

  let hasPdf = false;
  for (let i = 0; i < 5; i++) {
    const f = document.querySelector('[name="pdf_'+i+'"]');
    if (f && f.files && f.files[0]) { hasPdf = true; break; }
  }
  if (!hasPdf) {
    errorBox.textContent = 'Please upload at least one carrier PDF.';
    errorBox.classList.remove('d-none');
    return;
  }

  document.getElementById('overlay').style.display = 'flex';
  document.getElementById('generateBtn').disabled = true;

  try {
    const formData = new FormData(document.getElementById('quoteForm'));
    const response = await fetch('/generate', {method:'POST', body:formData});
    const ct = response.headers.get('Content-Type') || '';

    if (!response.ok || ct.includes('json')) {
      const text = await response.text();
      let msg = 'Server error ' + response.status;
      try { msg = JSON.parse(text).error || msg; } catch(e) { msg = text.substring(0,200); }
      throw new Error(msg);
    }

    const blob = await response.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    const cd   = response.headers.get('Content-Disposition') || '';
    const m    = cd.match(/filename="?([^"]+)"?/);
    a.download = m ? m[1] : 'Medical_Quote_Comparison.xlsx';
    a.href     = url;
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(url);

  } catch(err) {
    errorBox.textContent = 'Error: ' + err.message;
    errorBox.classList.remove('d-none');
    window.scrollTo(0,0);
  } finally {
    document.getElementById('overlay').style.display = 'none';
    document.getElementById('generateBtn').disabled = false;
  }
}
</script>
</body>
</html>"""


def extract_pdf_text(file_bytes):
    text_pages = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text_pages.append(t)
    return "\n\n".join(text_pages)


def parse_pdf_with_ai(pdf_text, plan_filter, enrollment):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable is not set.")
    client = anthropic.Anthropic(api_key=api_key)
    system = ("You are a health insurance data extraction specialist. "
              "You read carrier proposal PDFs and extract structured plan benefit data. "
              "You ALWAYS respond with valid JSON only — no explanation, no markdown, no code blocks. "
              "Monthly rates must be numbers (floats). All other values are strings. "
              "If a value is not found, use an empty string.")
    prompt = f"""Extract health insurance plan data from this carrier proposal.

FILTER: Only include plans where the individual in-network deductible matches: {plan_filter}
If filter is "all plans", include all plans found (max 3).

Return ONLY this JSON structure:
{{
  "carrier_name": "Carrier name",
  "commissions": "Commission info or empty string",
  "carrier_comments": "Funding type, network, PBM, quote ID, stop-loss, key notes",
  "plans": [
    {{
      "plan_name": "Full plan name",
      "network": "Network name and PBM",
      "deductible": "Ind / Fam in-network",
      "coinsurance": "Member % / Plan % description",
      "moop": "Ind / Fam OOPM in-network",
      "pcp": "PCP cost share",
      "telehealth": "Telehealth cost and vendor",
      "specialist": "Specialist cost share",
      "inpatient": "Inpatient hospitalization",
      "outpatient": "Outpatient facility",
      "er": "Emergency room",
      "urgent_care": "Urgent care",
      "lab": "Laboratory",
      "xray": "X-ray",
      "imaging": "CT/MRI/PET",
      "oon_ded": "OON deductible Ind / Fam",
      "oon_cost": "OON coinsurance",
      "oon_moop": "OON OOPM Ind / Fam",
      "rx_ded": "Rx deductible",
      "rx_copays": "Generic / Brand / Non-pref / Specialty",
      "rate_ee": 0.00,
      "rate_es": 0.00,
      "rate_ec": 0.00,
      "rate_ef": 0.00,
      "rate_guarantee": "12 Months"
    }}
  ]
}}

PDF TEXT:
{pdf_text[:14000]}"""

    message = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        system=system,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.rstrip("`").strip()
    return json.loads(raw)


def sv(ws, row, col, val, number_fmt=None):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == "MergedCell":
        return
    cell.value = val
    cell.protection = Protection(locked=False)
    if number_fmt:
        cell.number_format = number_fmt


def populate_excel(group_name, effective_date, quote_date, enrollment, carrier_groups):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb["Medical"]
    ws["C3"].value = group_name
    ws["C3"].protection = Protection(locked=False)
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=False, shrink_to_fit=False)
    sv(ws, 5, 9, f"Effective Date: {effective_date}   |   Quote Date: {quote_date}")
    for row, key in [(34,"ee"),(35,"es"),(36,"ec"),(37,"ef")]:
        for col in [4,5,6]:
            try:
                sv(ws, row, col, int(enrollment.get(key, 0)))
            except (ValueError, TypeError):
                sv(ws, row, col, 0)
    for gi, group in enumerate(carrier_groups):
        if not group:
            continue
        c1, c2, c3 = GROUP_COLS[gi]
        sv(ws, 9,  c1, group.get("carrier_name", ""))
        sv(ws, 53, c1, group.get("commissions", ""))
        sv(ws, 39, c1, group.get("carrier_comments", ""))
        for pi, plan in enumerate(group.get("plans", [])[:3]):
            col = [c1,c2,c3][pi]
            sv(ws, 10, col, plan.get("plan_name", ""))
            for field, row in ROW_MAP.items():
                val = plan.get(field, "")
                if field.startswith("rate_") and val:
                    try:
                        sv(ws, row, col, float(val), '"$"#,##0.00')
                    except (ValueError, TypeError):
                        sv(ws, row, col, val)
                else:
                    sv(ws, row, col, val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/generate", methods=["POST"])
def generate():
    try:
        group_name     = request.form.get("group_name", "Quote")
        effective_date = request.form.get("effective_date", "")
        quote_date     = request.form.get("quote_date", str(datetime.date.today()))
        plan_filter    = request.form.get("plan_filter", "all plans")
        enrollment = {
            "ee": request.form.get("enroll_ee", "0"),
            "es": request.form.get("enroll_es", "0"),
            "ec": request.form.get("enroll_ec", "0"),
            "ef": request.form.get("enroll_ef", "0"),
        }
        carrier_groups = []
        for gi in range(5):
            pdf_file = request.files.get(f"pdf_{gi}")
            if pdf_file and pdf_file.filename:
                pdf_bytes  = pdf_file.read()
                pdf_text   = extract_pdf_text(pdf_bytes)
                group_data = parse_pdf_with_ai(pdf_text, plan_filter, enrollment)
                carrier_groups.append(group_data)
            else:
                carrier_groups.append(None)
        if not any(carrier_groups):
            return jsonify({"error": "Please upload at least one carrier PDF."}), 400
        buf = populate_excel(group_name, effective_date, quote_date, enrollment, carrier_groups)
        safe_name = "".join(c for c in group_name if c.isalnum() or c in " _-").strip()
        filename  = f"{safe_name or 'Medical_Quote'}_Comparison.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
